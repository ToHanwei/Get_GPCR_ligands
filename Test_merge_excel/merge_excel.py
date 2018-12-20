#!codeing:utf-8

import sys
import os
import openpyxl

#result = "result.xlsx"
#wbResult = openpyxl.Workbook()
#wsResult = wbRessult.worksheets[0]
#wsResult.append(['a', 'b', 'c', 'd', 'e', 'f', 'g'])

def merge_cells(ws):
	rows = [''] + list(ws.rows)
	index1 = 2
	rowCount = len(rows)
	flag = "jump"
	while index1 < rowCount:
		value1 = rows[index1][0].value
		for index2, row2 in enumerate(rows[index1+1:], index1+1):
			value2 = row2[0].value
			if value1 == value2: continue
			ws.merge_cells("A"+str(index1)+":A"+str(index2-1))
			flag = "no"; break
		if flag == "jump":
			index2 = rowCount-1
			ws.merge_cells("A"+str(index1)+":A"+str(index2))
			break
		if index2 == rowCount-1:
			if rows[index2][0].value == rows[index2-1][0].value:
				ws.merge_cells("A"+str(index1)+":A"+str(index2))
				break
			else:
				ws.merge_cells("A"+str(index1)+":A"+str(index2-1))
				break
		index1 = index2


in_dir = sys.argv[1]
for infile in os.listdir(in_dir):
	print(infile)
	in_path = in_dir + "/" + infile
	wb = openpyxl.load_workbook(in_path)
	for ws in wb.worksheets:
		merge_cells(ws)
	wb.save(in_path)



#while index1 < 111:
#	value = rows[index1][0].value
#	for index2, row2 in enumerate(rows[index1+1:], index1+1):
#		print(row2[0].value == value)
#		if not (row2[0].value == value):
#			break
#		else:
#			ws.merge_cells("A"+str(index1)+":A"+str(index2-1))
#			break
#	ws.merge_cells("A"+str(index1)+":A"+str(index2-1))
#	print("A"+str(index1)+":A"+str(index2-1))
#	index1 = index2

